import wx
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import *
import os

class windowClass(wx.Frame):
    
    def __init__(self, *args, **kwargs):
        super(windowClass, self).__init__(*args, **kwargs, size=(400,670))

        self.basicGUI()

    def basicGUI(self):

        # Menu Bar
        panel = wx.Panel(self)
        menuBar = wx.MenuBar()
        fileButton = wx.Menu()
                
        infoItem = wx.MenuItem(fileButton, wx.ID_ANY, 'O programie\tCtrl+H')
        fileButton.Append(infoItem)

        ExitItem = wx.MenuItem(fileButton, wx.ID_EXIT, 'Zamknij\tCtrl+Q')
        fileButton.Append(ExitItem)

        toolBar = self.CreateToolBar()

        menuBar.Append(fileButton, 'Plik')
    
        self.SetMenuBar(menuBar)
        
        self.Bind(wx.EVT_MENU, self.info, infoItem)
        self.Bind(wx.EVT_MENU, self.Quit, ExitItem)
        # Menu Bar end
        
        def block_non_numbers(event):
            key_code = event.GetKeyCode()

            if ord('0') <= key_code <= ord('9'):        # Allow ASCII numerics
                event.Skip()
                return

            if key_code == ord('.'):        # Allow decimal points
                event.Skip()
                return

            if key_code == ord('\t'):       # Allow tabs, for tab navigation between TextCtrls
                event.Skip()
                return
            
            if key_code == ord('\b'):       # Allow backspace
                event.Skip()
                return
            
            return      # Block everything else

        def block_non_numbers1(event):      #Blocks also dots
            key_code = event.GetKeyCode()

            if ord('0') <= key_code <= ord('9'):        # Allow ASCII numerics
                event.Skip()
                return
            
            if key_code == ord('\t'):       # Allow tabs, for tab navigation between TextCtrls
                event.Skip()
                return
            
            if key_code == ord('\b'):       # Allow backspace
                event.Skip()
                return
            
            return      # Block everything else

        wx.StaticText(panel, -1, 'Temperatura:', (5,0))
        
        wx.StaticText(panel, -1, 'T1', (5,100))
        self.T1Txt = wx.TextCtrl(panel, pos=(25, 100), size=(50, 18))
        self.T1Txt.Bind(wx.EVT_CHAR, block_non_numbers)

        wx.StaticText(panel, -1, 'T2', (300,100))
        self.T2Txt = wx.TextCtrl(panel, pos=(320, 100), size=(50, 18))
        self.T2Txt.Bind(wx.EVT_CHAR, block_non_numbers)

        wx.StaticText(panel, -1, 'T3', (150,25))
        self.T3Txt = wx.TextCtrl(panel, pos=(170, 25), size=(50, 18))
        self.T3Txt.Bind(wx.EVT_CHAR, block_non_numbers)
        
        wx.StaticText(panel, -1, 'T4', (150,175))
        self.T4Txt = wx.TextCtrl(panel, pos=(170, 175), size=(50, 18))
        self.T4Txt.Bind(wx.EVT_CHAR, block_non_numbers)

        plate = wx.TextCtrl(panel, pos=(90, 60), size=(200, 100))
        plate.Enable(False)
        
        wx.StaticText(panel, -1, 'Ilość komórek siatki numerycznej:', (5,200))

        wx.StaticText(panel, -1, 'x', (5,225))
        self.ndsxTxt = wx.TextCtrl(panel, pos=(25, 225), size=(50, 18))
        self.ndsxTxt.Bind(wx.EVT_CHAR, block_non_numbers1)
        
        wx.StaticText(panel, -1, 'y', (150,225))
        self.ndsyTxt = wx.TextCtrl(panel, pos=(170, 225), size=(50, 18))
        self.ndsyTxt.Bind(wx.EVT_CHAR, block_non_numbers1)

        wx.StaticText(panel, -1, 'Wymiary ścianki [m]:', (5,250))

        wx.StaticText(panel, -1, 'x', (5,275))
        self.lxTxt = wx.TextCtrl(panel, pos=(25, 275), size=(50, 18))
        self.lxTxt.Bind(wx.EVT_CHAR, block_non_numbers)
        
        wx.StaticText(panel, -1, 'y', (5,300))
        self.lyTxt = wx.TextCtrl(panel, pos=(25, 300), size=(50, 18))
        self.lyTxt.Bind(wx.EVT_CHAR, block_non_numbers)

        precList = ['Półprecyzja', 'Precyzja pojedyncza', 'Precyzja podwójna'] 
        self.precBox = wx.RadioBox(panel, label = 'Precyzja zapisu liczb', pos = (5,325), choices = precList,
            majorDimension = 3, style = wx.RA_SPECIFY_ROWS)

        methList = ['Jacobi', 'Gauss-Seidl', 'SOR']     
        self.methBox = wx.RadioBox(panel, label = 'Metoda obliczeń', pos = (150,325), choices = methList,
            majorDimension = 3, style = wx.RA_SPECIFY_ROWS)

        wx.StaticText(panel, -1, 'Dodatkowe parametry', (5,425))

        wx.StaticText(panel, -1, 'Dokładność obliczeń:', (5,450))
        wx.StaticText(panel, -1, '10', (140,450))
        wx.StaticText(panel, -1, '-', (153,445))
        self.epsTxt = wx.TextCtrl(panel, pos=(160, 445), size=(20, 18))
        self.epsTxt.Bind(wx.EVT_CHAR, block_non_numbers1)

        wx.StaticText(panel, -1, 'Maksymalna ilość iteracji:', (5,475))
        self.iteracjeTxt = wx.TextCtrl(panel, pos=(160, 475), size=(50, 18))
        self.iteracjeTxt.Bind(wx.EVT_CHAR, block_non_numbers1)

        wx.StaticText(panel, -1, 'Współczynnik relaksacji "\u03C9":', (5,500))
        self.relaksacjaTxt = wx.TextCtrl(panel, pos=(160, 500), size=(50, 18))
        self.relaksacjaTxt.Bind(wx.EVT_CHAR, block_non_numbers)    

        self.obliczBtn = wx.Button(panel, label = 'Oblicz!', pos = (5, 525), size = (100,50))
        self.Bind(wx.EVT_BUTTON, self.oblicz, self.obliczBtn)
       
        self.SetTitle('Solver')
        self.Show(True)

    def Quit (self, e):
        self.Close()

    def info (self, e):

        alert = wx.MessageDialog(None, 'Instrukcja obsługi programu \n\
\n\
Program służy do obliczania rozkładu temperatury na ściance płaskiej.\n\
Do wprowadzenia wartości w postaci ułamków dziesiętnych należy użyć kropki.\n\
Wartości ilości komórek siatki numerycznej, dokładności obliczeń i maksymalnej ilości iteracji\n\
można wprowadzić wyłącznie jako liczby całkowite.\n\
1. Wpisz wartości temperatury (T1, T2, T3, T4) na ściankach.\n\
2. Wpisz ilość komórek siatki numerycznej wzdłuż boków ścianki.\n\
3. Wprowadź wymiary ścianki prostokątnej.\n\
4. Wybierz jeden z trzech wariantów precyzji obliczeń numerycznych.\n\
5. Wybierz jedną z trzech metod obliczeń.\n\
6. Wprowadź żądaną dokładność obliczeń oraz maksymalną ilość iteracji, jaką ma wykonać program.\n\
7. Jeśli wybrałeś metodę SOR, wprowadź wartość współczynnika relaksacji.\n\
Wartość współczynnika relaksacji powinna zawierać się w przedziale (0,2).\n\
\n\
Pamiętaj, że jeśli zadasz zbyt dużą dokładność, ilość iteracji oraz ilość komórek siatki\n\
numerycznej wzdłuż boków płyty program będzie wykonywał obliczenia bardzo długo.\n\n\
Gdy program nie odpowiada, to znaczy, że w danym momencie wykonuje obliczenia.\n\n\
Po zakończonych obliczeniach program wygeneruje wykres, który można zapisać w formie\n\
graficznej, a po jego zamknięciu wyeksportuje także wyniki do pliku arkusza\n\
kalkulacyjnego o rozszerzeniu .xlsx.', 'O programie', wx.OK)
        alert.ShowModal()
        

    def oblicz (self, event):
        
        self.T1 = float(self.T1Txt.GetValue())
        self.T2 = float(self.T2Txt.GetValue())
        self.T3 = float(self.T3Txt.GetValue())
        self.T4 = float(self.T4Txt.GetValue())
        self.lx = float(self.lxTxt.GetValue())
        self.ly = float(self.lyTxt.GetValue())
        self.ndsx = int(self.ndsxTxt.GetValue())
        self.ndsy = int(self.ndsyTxt.GetValue())
        self.prec = self.precBox.GetSelection()
        self.method = self.methBox.GetSelection()
        self.eps = 10**-(int(self.epsTxt.GetValue()))
        self.iteracje = int(self.iteracjeTxt.GetValue())

        T1 = self.T1
        T2 = self.T2
        T3 = self.T3
        T4 = self.T4
        lx = self.lx
        ly = self.ly
        ndsx = self.ndsx
        ndsy = self.ndsy
        prec = self.prec
        method = self.method
        eps = self.eps
        iteracje = self.iteracje
                
        dx = lx/ndsx
        dy = ly/ndsy
        
        def solver(T1, T2, T3, T4, ndsx, ndsy, dx, dy, prec):
    
            T = np.ones((ndsy+1,ndsx+1)) * (T1+T2+T3+T4)/4
            if prec == 0:        # half-precision
                T = np.float16(T)
            elif prec == 1:     # single-precision
                T = np.float32(T)
            elif prec == 2:     # double-precision
                T = np.float64(T)

            T[:,0] = T1
            T[:,-1] = T2
            T[0,:] = T3
            T[-1,:] = T4
            T[0,0] = (T1+T3)/2
            T[0,-1] = (T2+T3)/2
            T[-1,0] = (T1+T4)/2
            T[-1,-1] = (T3+T4)/2
            
            TOld = T*1
            mian = 2/(dx**2)+2/(dy**2)
            h = 0       # number of loop iterations
            delta = 1
            
            if method == 0:     # MRS
                while delta > eps:
                    h = h + 1
                    for i in range(1, ndsy):
                        for j in range (1, ndsx):
                            T[i,j] = ((TOld[i+1,j]+TOld[i-1,j])/(dx**2)+(TOld[i,j+1]+TOld[i,j-1])/(dy**2))/mian                    
                    delta = np.amax(abs(T-TOld))
                    TOld = T*1
                    print('iteracja =', h)
                    if h >= iteracje:
                        break

            elif method == 1:       # Gauss-Seidl
                while delta > eps:
                    h = h + 1
                    for i in range(1, ndsy):
                        for j in range (1, ndsx):
                            T[i,j] = ((TOld[i+1,j]+T[i-1,j])/(dx**2)+(TOld[i,j+1]+T[i,j-1])/(dy**2))/mian
                    delta = np.amax(abs(T-TOld))
                    TOld = T*1
                    print('iteracja =', h)
                    if h >= iteracje:
                        break

            elif method == 2:       # SOR
                self.w = float(self.relaksacjaTxt.GetValue())
                w = self.w                
                while delta > eps:
                    h = h + 1
                    for i in range(1, ndsy):
                        for j in range (1, ndsx):
                            T[i,j] = (1-w)*TOld[i,j]+w*((TOld[i+1,j]+T[i-1,j])/(dx**2)+(TOld[i,j+1]+T[i,j-1])/(dy**2))/mian
                    delta = np.amax(abs(T-TOld))
                    TOld = T*1
                    print('iteracja =', h)
                    if h >= iteracje:
                        break
            
            self.h = h
            self.T = T
            return T

        def visualize(T, ndsx, ndsy, lx, ly, dx, dy):

            X, Y = np.meshgrid(np.arange(0, lx+dx/2, dx), np.arange(0, ly+dy/2, dy))
            plt.contourf(X, Y, T, cmap=plt.cm.jet)
            plt.colorbar(label='Temperatura [\u2103]')
            plt.show()

        def xlsxExport(self,event):

            self.wb = Workbook()
            ws = self.wb.active

            ws.title = 'Temperatura'

            ws['A1'] = 'T1'
            ws['A2'] = 'T2'
            ws['A3'] = 'T3'
            ws['A4'] = 'T4'
            ws['A5'] = 'Komórki x'
            ws['A6'] = 'Komórki y'
            ws['A7'] = 'lx'
            ws['A8'] = 'ly'
            ws['A9'] = 'Precyzja'
            ws['A10'] = 'Dokładność'
            ws['A11'] = 'Iteracje'

            ws['B1'] = self.T1
            ws['B2'] = self.T2
            ws['B3'] = self.T3
            ws['B4'] = self.T4
            ws['B5'] = self.ndsx
            ws['B6'] = self.ndsy
            ws['B7'] = self.lx
            ws['B8'] = self.ly
            ws['B9'] = self.prec
            ws['B10'] = self.eps
            ws['B11'] = self.h

            for i in range (0, ndsx+1):
                for j in range (0, ndsy+1):
                    ws.cell(row = j+4, column = i+4).value = T[j,i]
            
            def AcceptNameEx(event):
            
                excelGenerator = 0

                rawName = self.excelFilename.GetLineText(0)
                counter = rawName.count("#")
                counter = counter + rawName.count("%") 
                counter = counter + rawName.count("&") 
                counter = counter + rawName.count("*") 
                counter = counter + rawName.count(":")
                counter = counter + rawName.count("?")
                counter = counter + rawName.count("/")
                counter = counter + rawName.count("\\")
                counter = counter + rawName.count("|") 

                filenameEx = rawName + ".xlsx"
                self.osPath = os.getcwd()     
                pathTable = self.osPath[0].split('\\')
                
                self.path = ""
                for i in range (0, len(pathTable)-1):
                    self.path = self.path + pathTable[i] + "\\"
                self.path = self.path + filenameEx

                if os.path.isfile(self.path) == 0:
                    excelGenerator = 1
                else:
                    fileExists = wx.MessageDialog(self.windowExName, "Wrong file name. \
File with that name exists. Delete existing file or choose another name.",\
                                                    "Wrong filename", style = wx.OK_DEFAULT | wx.ICON_ERROR | wx.CENTRE)
                    fileExists.ShowModal()
                    self.windowExName.Show(True)

                if counter == 0 and excelGenerator == 1:
                    self.wb.save(filenameEx)
                    fileGenerated = wx.MessageDialog(self.windowExName, "Excel file \
generated successfully in: " + self.path, "File saved",\
                                                        style = wx.OK_DEFAULT | wx.ICON_INFORMATION | wx.CENTRE)
                    if fileGenerated.ShowModal() == wx.ID_OK:
                        self.windowExName.Destroy()
                elif counter != 0:
                    nameError = wx.MessageDialog(self.windowExName, "Wrong file name. \
It cannot contain '#','%','&','*',':','?','/','\\','|'.", "Wrong filename",\
                                                 style = wx.OK_DEFAULT | wx.ICON_ERROR | wx.CENTRE)
                    nameError.ShowModal()
                    self.windowExName.Show(True)

            self.windowExName = wx.Dialog(self, title = 'Choose .xlsx filename',\
                                          size = (150, 130), style = wx.STAY_ON_TOP | wx.MAXIMIZE_BOX | wx.MINIMIZE_BOX | wx.CAPTION | wx.CLOSE_BOX)
            self.filenameInfoEx = wx.StaticText(self.windowExName, wx.ID_ANY,\
                                                label = 'Set the .xlsx filename:', pos = (10, 5))
            self.excelFilename = wx.TextCtrl(self.windowExName, value = "",\
                                             name = 'excelFilename', pos = (0,25), size = (140, 25), style = wx.TE_PROCESS_ENTER)
            self.acceptFilenameEx = wx.Button(self.windowExName, wx.ID_ANY, 'Ok', (45,60), (50,25))
            self.acceptFilenameEx.Bind(wx.EVT_BUTTON, AcceptNameEx, self.acceptFilenameEx)

            self.windowExName.Show()

        T = solver(T1, T2, T3, T4, ndsx, ndsy, dx, dy, prec)
        visualize(T, ndsx, ndsy, lx, ly, dx, dy)
        xlsxExport(self, event)

def main():
    app = wx.App()
    windowClass(None)
    
    app.MainLoop()
    
main()
